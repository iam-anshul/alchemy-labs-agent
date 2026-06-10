"""Multi-stage reasoning agent over the document tree.

Pipeline:

    answer_query(workspace_id, user_id, query, doc_ids=None)
        │
        ├── 1. Router agent (high recall)
        │      walks tree, returns:
        │        - page_targets:  [(doc_id, p_start, p_end, reason), ...]
        │        - table_targets: [(doc_id, table_id, reason), ...]
        │
        ├── 2. Excel agent (only if table_targets exist)
        │      loads tables as pandas DataFrames, runs pandas via tool,
        │      returns structured `TableFinding`s.
        │
        └── 3. Answer agent
               grounds final answer on rendered pages + Excel findings,
               returns AnswerResult with citations + confidence.

All three are pydantic-ai Agents talking to the OpenAI-compatible endpoint
configured in `.env`.
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import time
import uuid
from pathlib import Path
from collections import deque
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from pydantic_ai import Agent, RunContext
from pydantic_ai.models.openai import OpenAIChatModel, OpenAIChatModelSettings
from pydantic_ai.providers.openai import OpenAIProvider
from pydantic_ai.usage import UsageLimits
from sqlalchemy import select

from agent_schemas import (
    AnswerResult,
    Citation,
    ExcelResult,
    HopTrace,
    PageTarget,
    QueryAnswer,
    RouterResult,
    TableFinding,
    TableTarget,
)
from api.events import EventSink, file_artifact
from config import get_settings
from db import SessionLocal, utils
from db.models import Doc, ExtractedTable, Node, Page

log = logging.getLogger(__name__)


# ===========================================================================
# Model factory
# ===========================================================================

_cached_model: OpenAIChatModel | None = None


def _build_model() -> OpenAIChatModel:
    """Return a shared OpenAIChatModel instance (created once, reused)."""
    global _cached_model
    if _cached_model is not None:
        return _cached_model
    s = get_settings()
    if not s.openai_api_key:
        raise RuntimeError("OPENAI_API_KEY is not set in .env")
    provider = OpenAIProvider(api_key=s.openai_api_key, base_url=s.openai_base_url)
    settings = OpenAIChatModelSettings(extra_body={"enable_thinking": False})
    _cached_model = OpenAIChatModel(s.openai_model, provider=provider, settings=settings)
    return _cached_model


# ===========================================================================
# Router agent
# ===========================================================================

@dataclass
class RouterDeps:
    workspace_id: str
    user_id: str
    candidate_doc_ids: list[str]
    # Hard guard: refuse to re-expand or expand-leaf — defends against the LLM
    # ignoring system prompt instructions.
    _expanded_nodes: set[str] = field(default_factory=set)
    _known_leaves: set[str] = field(default_factory=set)
    # Multi-hop dedup: page ranges and tables already emitted in previous hops.
    already_picked_pages: set[str] = field(default_factory=set)   # "{doc_id}:{start}-{end}"
    already_picked_tables: set[str] = field(default_factory=set)  # table_id


_ROUTER_SYSTEM = """\
You are a document routing agent. Your job is to pick the SPECIFIC pieces of evidence \
that should be read to answer the user's query — across one or more documents.

You will see:
- The user's query.
- A list of candidate documents with their top-level summaries.

Tools you can use (call as many as needed, in any order):
- `expand_doc(doc_id)`: top-level sections of a document.
- `expand_node(node_id)`: children of a tree node, with brief summaries.
- `peek_node(node_id)`: full summary of a node (use when brief is not enough).
- `list_doc_tables(doc_id)`: every extracted table in a doc — id, source page, row count, column names if known.
- `peek_table(table_id)`: detailed schema + first few rows of one table.

CRITICAL BUDGET: you have a HARD LIMIT of ~30 total tool calls across the whole task. \
Stay well under it. Be DELIBERATE — every tool call is expensive.

Routing strategy:

1. **Read the doc summaries already provided in the user message.** Decide which docs are \
plausibly relevant. Skip docs that are clearly off-topic.

2. For each plausible doc, call `expand_doc` ONCE. You will get back its top-level sections \
with brief summaries. Each section also tells you `has_children` and `pages` (its page range).

3. For each section, pick exactly ONE of three actions based on its brief summary:
     (a) **skip** — clearly irrelevant
     (b) **take** — relevant; emit it as a page_target directly (no further drilling)
     (c) **drill** — relevant but its page range is wide AND its summary is too vague

Strongly prefer (a) or (b). Only drill (c) when truly necessary.

4. **NEVER call `expand_node` on a node whose brief shows `has_children: false`.** \
Those nodes are leaves — drilling returns nothing. Use them as page_targets directly.

5. **NEVER call `expand_node` more than 2 levels deep from the root.** Stop and commit.

6. **STOP DRILLING as soon as a node's range is ≤ ~25 pages and its summary matches the query.** \
A depth-1 or depth-2 node is a perfectly fine target. You do not need to find the smallest leaf.

7. Use `peek_node(node_id)` ONLY when a brief summary is genuinely ambiguous (rare).

8. **Tables — only when the query is numeric / aggregation / ranking / comparison / "top N" / \
"how much"**: call `list_doc_tables` for the 1-2 most relevant docs, then `peek_table` on the \
1-3 whose columns clearly match. Add them to `table_targets`. Do NOT enumerate all tables.

9. Emit `page_targets` for narrative answers, `table_targets` for tabular answers. \
A query may need both. Aim for **3-6 total targets**, NOT 20+.

Rules:
- Do NOT try to answer the query yourself — emit targets only.
- Each target must include a one-line `reason`.
- If absolutely nothing is relevant, return empty lists and explain in `reasoning`.
- Recall matters, but so does focus — over-exploring is just as bad as missing.
"""


def _brief_node(n: Node) -> dict[str, Any]:
    summary = (n.summary or "")
    if len(summary) > 1200:
        summary = summary[:1200] + " …"
    return {
        "node_id": n.node_id,
        "title": n.title,
        "pages": f"{n.start_page}-{n.end_page}",
        "depth": n.depth,
        "summary": summary,
        "n_tables": len(n.table_ids or []),
        "has_children": bool(n.child_ids),
    }


def _brief_table(t: ExtractedTable, columns: list[str] | None = None) -> dict[str, Any]:
    return {
        "table_id": t.table_id,
        "doc_id": t.doc_id,
        "source_page": t.source_page,
        "title": t.title_guess,
        "description": t.description,
        "row_count": t.row_count,
        "columns": columns or [],
    }


def _build_router(
    *,
    system_prompt: str = _ROUTER_SYSTEM,
) -> Agent[RouterDeps, RouterResult]:
    agent: Agent[RouterDeps, RouterResult] = Agent(
        _build_model(),
        deps_type=RouterDeps,
        output_type=RouterResult,
        system_prompt=system_prompt,
        retries=2,
    )

    @agent.tool
    async def expand_doc(ctx: RunContext[RouterDeps], doc_id: str) -> list[dict[str, Any]]:
        """Return top-level (depth-1) sections of a document.

        Args:
            doc_id: Document identifier (must be in the candidate list).
        """
        log.info("[router] expand_doc(%s)", doc_id)
        if doc_id not in ctx.deps.candidate_doc_ids:
            log.warning("[router] expand_doc: %s not in candidates", doc_id)
            return [{"error": f"doc_id {doc_id} is not in the candidate list"}]
        with SessionLocal() as db:
            roots = utils.get_root_nodes(db, doc_id)
            if not roots:
                log.info("[router] expand_doc(%s) -> no roots", doc_id)
                return []
            root = roots[0]
            children = [db.get(Node, cid) for cid in (root.child_ids or [])]
            result = [_brief_node(c) for c in children if c]
            log.info("[router] expand_doc(%s) -> %d section(s): %s",
                     doc_id, len(result), [r["title"] for r in result])
            return result

    @agent.tool
    async def expand_node(ctx: RunContext[RouterDeps], node_id: str) -> list[dict[str, Any]]:
        """Return the children of a tree node.

        Args:
            node_id: Parent node identifier.
        """
        log.info("[router] expand_node(%s)", node_id)

        # Hard guards against repeated / pointless calls
        if node_id in ctx.deps._known_leaves:
            log.warning("[router] expand_node: %s is a known leaf — REJECTED", node_id)
            return [{
                "error": "REJECTED — this node was already identified as a leaf earlier. "
                         "Stop calling expand_node on it. Use it directly as a page_target."
            }]
        if node_id in ctx.deps._expanded_nodes:
            log.warning("[router] expand_node: %s was already expanded — REJECTED", node_id)
            return [{
                "error": "REJECTED — this node was already expanded earlier in this run. "
                         "Use the children you already received, or commit to targets."
            }]

        with SessionLocal() as db:
            node = db.get(Node, node_id)
            if node is None:
                log.warning("[router] expand_node: %s not found", node_id)
                return [{"error": f"node {node_id} not found"}]
            if node.doc_id not in ctx.deps.candidate_doc_ids:
                return [{"error": "node belongs to a non-candidate doc"}]
            if not node.child_ids:
                ctx.deps._known_leaves.add(node_id)
                log.info("[router] expand_node(%s) -> leaf (no children)", node_id)
                return [{
                    "note": (
                        f"Node {node_id} is a leaf — has_children=False. "
                        f"DO NOT call expand_node on it. "
                        f"Use it directly as a page_target: pages {node.start_page}-{node.end_page}."
                    )
                }]
            ctx.deps._expanded_nodes.add(node_id)
            children = [db.get(Node, cid) for cid in node.child_ids]
            # Pre-mark known leaves so the model sees has_children clearly
            for c in children:
                if c and not c.child_ids:
                    ctx.deps._known_leaves.add(c.node_id)
            result = [_brief_node(c) for c in children if c]
            log.info("[router] expand_node(%s) -> %d child(ren): %s",
                     node_id, len(result), [r["title"] for r in result])
            return result

    @agent.tool
    async def peek_node(ctx: RunContext[RouterDeps], node_id: str) -> dict[str, Any]:
        """Return the full summary of a single node (use when brief isn't enough).

        Args:
            node_id: Node identifier.
        """
        log.info("[router] peek_node(%s)", node_id)
        with SessionLocal() as db:
            node = db.get(Node, node_id)
            if node is None:
                return {"error": f"node {node_id} not found"}
            if node.doc_id not in ctx.deps.candidate_doc_ids:
                return {"error": "node belongs to a non-candidate doc"}
            log.info("[router] peek_node(%s) -> %r p%s-p%s",
                     node_id, node.title, node.start_page, node.end_page)
            return {
                "node_id": node.node_id,
                "doc_id": node.doc_id,
                "title": node.title,
                "start_page": node.start_page,
                "end_page": node.end_page,
                "depth": node.depth,
                "summary": node.summary,
                "table_ids": node.table_ids or [],
                "has_children": bool(node.child_ids),
            }

    @agent.tool
    async def list_doc_tables(ctx: RunContext[RouterDeps], doc_id: str) -> list[dict[str, Any]]:
        """List all extracted tables in a document with id, source page, row count, and columns.

        Use this when the query needs tabular data (aggregation, filtering, ranking, lookups).

        Args:
            doc_id: Document identifier.
        """
        log.info("[router] list_doc_tables(%s)", doc_id)
        if doc_id not in ctx.deps.candidate_doc_ids:
            return [{"error": f"doc_id {doc_id} is not in the candidate list"}]
        with SessionLocal() as db:
            tables = utils.list_tables_for_doc(db, doc_id)
            out: list[dict[str, Any]] = []
            for t in tables:
                cols = _peek_xlsx_columns(t.xlsx_bytes) if t.xlsx_bytes else []
                out.append(_brief_table(t, cols))
            log.info("[router] list_doc_tables(%s) -> %d table(s)", doc_id, len(out))
            return out

    @agent.tool
    async def peek_table(ctx: RunContext[RouterDeps], table_id: str) -> dict[str, Any]:
        """Return schema + first few rows of one extracted table.

        Args:
            table_id: Table identifier.
        """
        log.info("[router] peek_table(%s)", table_id)
        with SessionLocal() as db:
            t = db.get(ExtractedTable, table_id)
            if t is None:
                return {"error": f"table {table_id} not found"}
            if t.doc_id not in ctx.deps.candidate_doc_ids:
                return {"error": "table belongs to a non-candidate doc"}
            preview = _peek_xlsx_rows(t.xlsx_bytes, max_rows=6) if t.xlsx_bytes else []
            cols = preview[0] if preview else []
            log.info("[router] peek_table(%s) -> page=%s cols=%s rows=%s",
                     table_id, t.source_page, cols, t.row_count)
            return {
                "table_id": t.table_id,
                "doc_id": t.doc_id,
                "source_page": t.source_page,
                "title": t.title_guess,
                "description": t.description,
                "row_count": t.row_count,
                "columns": cols,
                "preview_rows": preview[1:6],
            }

    return agent


# ===========================================================================
# Excel / table agent
# ===========================================================================

@dataclass
class ExcelDeps:
    # {table_id: DataFrame}. The agent can reference these by table_id via tools.
    dfs: dict[str, pd.DataFrame] = field(default_factory=dict)
    # {table_id: doc_id} for citation back-tracking
    doc_id_by_table: dict[str, str] = field(default_factory=dict)


_EXCEL_SYSTEM = """\
You are a tabular-data analyst. The user asked a question, and the router selected one or more \
tables (already loaded as pandas DataFrames) that may contain the answer.

Tools:
- `describe_table(table_id)`: returns columns, dtypes, shape, and first 5 rows.
- `run_pandas(table_id, code)`: evaluates ONE pandas EXPRESSION against `df` (the requested table). \
The expression must be evaluable — NOT a statement.

Examples that WORK:
  - `df['Revenue'].sum()`
  - `df[df['Country']=='India']['Sales'].max()`
  - `df.groupby('Segment')['Profit'].sum().to_dict()`
  - `df['Amount'].str.replace(',', '').astype(float).sum()`   # for numbers stored as strings with commas
  - `int(df.iloc[0]['Total'].replace(',', ''))`               # type casts are fine

Examples that FAIL (do NOT use):
  - `x = df[...].sum()`            # assignment → not an expression
  - `import pandas as pd`          # imports forbidden
  - `df.to_csv('out.csv')`         # I/O forbidden

Available built-ins inside the sandbox: int, float, str, bool, len, sum, min, max, abs, \
round, sorted, reversed, list, dict, tuple, set, range, enumerate, zip, map, filter, any, all.

Process:
1. Call `describe_table` first to confirm column names and types. Numeric-looking columns \
are often stored as STRINGS with commas (e.g. "67,111") — strip commas and cast to float/int.
2. Use `run_pandas` to compute what the query needs. Keep expressions small. Batch related \
metrics into ONE expression where possible (`df.pct_change`, `groupby`, etc.) — pandas is vectorized.
3. Return one `TableFinding` per relevant table with precise numbers and units.

Rules:
- Do NOT invent column names — only use ones from `describe_table`.
- Pick 2–4 MOST query-relevant metrics per table. Don't enumerate every value.
- If a table isn't relevant after inspecting, omit it from `findings`.
- Use `notes` for caveats (missing values, unclear units, multi-table joins skipped).
- Report numbers in NATIVE units. Indian crore = 10,000,000; lakh = 100,000. Expand explicitly: \
"₹1,068.03 crore = ₹10,680,300,000".
"""


import builtins as _builtins_mod

_SAFE_BUILTINS: dict[str, Any] = {
    name: getattr(_builtins_mod, name)
    for name in (
        "int", "float", "str", "bool", "len", "sum", "min", "max",
        "abs", "round", "sorted", "reversed", "list", "dict", "tuple", "set",
        "range", "enumerate", "zip", "map", "filter", "any", "all",
        "isinstance", "type", "repr", "print",
    )
}


def _build_excel_agent() -> Agent[ExcelDeps, ExcelResult]:
    agent: Agent[ExcelDeps, ExcelResult] = Agent(
        _build_model(),
        deps_type=ExcelDeps,
        output_type=ExcelResult,
        system_prompt=_EXCEL_SYSTEM,
        retries=2,
    )

    @agent.tool
    async def describe_table(ctx: RunContext[ExcelDeps], table_id: str) -> dict[str, Any]:
        """Return columns, dtypes, shape, and head of one table.

        Args:
            table_id: Identifier of one of the tables loaded for this query.
        """
        log.info("[excel] describe_table(%s)", table_id)
        df = ctx.deps.dfs.get(table_id)
        if df is None:
            log.warning("[excel] describe_table: %s not loaded", table_id)
            return {"error": f"table {table_id} not loaded for this query"}
        log.info("[excel] describe_table(%s) -> shape=%s cols=%s",
                 table_id, list(df.shape), list(df.columns.astype(str)))
        return {
            "table_id": table_id,
            "shape": list(df.shape),
            "columns": list(df.columns.astype(str)),
            "dtypes": {str(k): str(v) for k, v in df.dtypes.items()},
            "head": df.head(5).astype(str).to_dict(orient="records"),
        }

    @agent.tool
    async def run_pandas(ctx: RunContext[ExcelDeps], table_id: str, code: str) -> dict[str, Any]:
        """Evaluate a pandas EXPRESSION (not a statement) against `df`, which is the requested table.

        Examples of valid `code`:
          - "df.shape"
          - "df['Revenue'].sum()"
          - "df.groupby('Segment')['Profit'].sum().head(10).to_dict()"

        Args:
            table_id: Table to run against.
            code: A single pandas expression. No imports, no assignments, no I/O.
        """
        log.info("[excel] run_pandas(%s) code=%r", table_id, code)
        df = ctx.deps.dfs.get(table_id)
        if df is None:
            return {"error": f"table {table_id} not loaded for this query"}
        forbidden = ("import ", "open(", "exec(", "eval(", "__")
        if any(f in code for f in forbidden):
            log.warning("[excel] run_pandas: forbidden token in code: %r", code)
            return {"error": "forbidden token in code"}
        try:
            result = eval(
                code,
                {"pd": pd, "df": df, "__builtins__": _SAFE_BUILTINS},
                {},
            )
        except SyntaxError:
            log.warning("[excel] run_pandas: not a valid expression (statement?): %r", code)
            return {"error": "code must be a single expression, not a statement. Avoid assignments like 'x = ...'"}
        except Exception as e:
            log.warning("[excel] run_pandas error: %s: %s", type(e).__name__, e)
            return {"error": f"{type(e).__name__}: {e}"}

        if isinstance(result, pd.DataFrame):
            preview = result.head(30).astype(str).to_dict(orient="records")
            log.info("[excel] run_pandas -> DataFrame(shape=%s, %d rows preview)",
                     list(result.shape), len(preview))
            return {"result": preview}
        if isinstance(result, pd.Series):
            preview = result.head(30).astype(str).to_dict()
            log.info("[excel] run_pandas -> Series(len=%d, %d entries preview)",
                     len(result), len(preview))
            return {"result": preview}
        result_str = str(result)
        log.info("[excel] run_pandas -> scalar=%s",
                 result_str if len(result_str) < 200 else result_str[:200] + "…")
        return {"result": result_str}

    return agent


# ===========================================================================
# Answer agent (final synthesiser)
# ===========================================================================

_ANSWER_SYSTEM = """\
You are a document-grounded answer agent. Answer the user's query using ONLY the content provided:
- Page extracts from one or more documents.
- Optional structured findings from a tabular analyst that already ran pandas on the relevant tables.

Rules:
- Quote / paraphrase from the provided content. Do not invent.
- The "Tabular analyst findings" block contains facts ALREADY EXTRACTED from the same \
documents. Treat findings as FIRST-CLASS EVIDENCE — equivalent to direct page content. \
They are NOT metadata about what is or isn't available.
- If the query needs arithmetic across multiple facts (one from pages, one from findings, \
or two findings), perform it. Do NOT refuse just because the inputs came from different blocks.
- If content does not contain the answer, say so explicitly and set confidence='low'.
- Set confidence='medium' for partial answers; 'high' only when fully grounded.
- Always include citations: doc_title + page range.
- Be concise but complete.

UNITS / CURRENCY:
- Keep figures in NATIVE units (₹ crore, $ million, etc.). Indian crore = 10,000,000; lakh = 100,000.
- If a `TableFinding` already contains a computed number, quote it verbatim. Don't recompute.
- Indian crore = 10,000,000; lakh = 100,000. Expand to absolute numbers when comparing units.
- If a calculation needs an external assumption (FX rate, battery size, etc.), state it explicitly \
and flag the assumption in your answer.

Follow-up control (used by an orchestrator to re-route the query if needed):
- Set `needs_more=True` ONLY when confidence is medium/low AND you can name SPECIFIC pieces \
of information that, if found in the documents, would let you raise confidence.
- Provide 1–3 entries in `follow_up_questions`. Each must be a precise SEARCH TARGET:
    GOOD: "What is HDFC Bank's FY25 annual CSR spend?"
    GOOD: "What was the cumulative PacifiCorp wildfire accrual as of Dec 31, 2024?"
    BAD:  "more context"
    BAD:  "everything about HDFC"
- Each follow-up will be processed as an independent investigation in the next hop.
- Set `needs_more=False` (and leave `follow_up_questions` empty) if the documents simply \
don't contain the answer OR if you are already at high confidence.

File saving:
- If the user's query explicitly asks to save / write / export / "as a note" /
  "save as <name>", set `save_to_file=True` and a kebab-case `suggested_filename`
  ending in `.md` (e.g. "hdfc-csr-summary-fy25.md").
- Otherwise leave both at defaults. Do NOT save unsolicited.
"""


def _build_answer_agent() -> Agent[None, AnswerResult]:
    return Agent(
        _build_model(),
        output_type=AnswerResult,
        system_prompt=_ANSWER_SYSTEM,
        retries=2,
    )


# ===========================================================================
# Helpers: xlsx / context rendering
# ===========================================================================

def _peek_xlsx_rows(xlsx_bytes: bytes | None, max_rows: int = 6) -> list[list[Any]]:
    """Read first max_rows rows (including header) from xlsx_bytes."""
    if not xlsx_bytes:
        return []
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        out: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            out.append(["" if v is None else v for v in row])
        return out
    except Exception:
        return []


def _peek_xlsx_columns(xlsx_bytes: bytes | None) -> list[str]:
    rows = _peek_xlsx_rows(xlsx_bytes, max_rows=1)
    if not rows:
        return []
    return [str(c) for c in rows[0]]


def _xlsx_to_dataframe(xlsx_bytes: bytes) -> pd.DataFrame:
    """Load xlsx bytes into a pandas DataFrame (first sheet)."""
    return pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=0)


def _xlsx_to_markdown(xlsx_bytes: bytes, max_rows: int = 40) -> str:
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as e:
        return f"_(could not read xlsx: {e})_"
    out: list[str] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))[:max_rows]
        if not rows:
            continue
        width = max(len(r) for r in rows)
        rows = [list(r) + [""] * (width - len(r)) for r in rows]
        header = rows[0]
        out.append("| " + " | ".join("" if v is None else str(v) for v in header) + " |")
        out.append("|" + "|".join("---" for _ in header) + "|")
        for row in rows[1:]:
            out.append("| " + " | ".join("" if v is None else str(v) for v in row) + " |")
        out.append("")
    return "\n".join(out).strip() or "_(empty table)_"


def _render_page_context(
    targets: list[PageTarget],
) -> tuple[str, list[Citation], list[str]]:
    """Render selected page ranges (with inline tables) for the answer agent."""
    blocks: list[str] = []
    citations: list[Citation] = []
    inline_table_ids: list[str] = []

    with SessionLocal() as db:
        for t in targets:
            doc = utils.get_doc(db, t.doc_id)
            if doc is None:
                continue
            pages = list(db.scalars(
                select(Page)
                .where(
                    Page.doc_id == t.doc_id,
                    Page.page_n >= t.start_page,
                    Page.page_n <= t.end_page,
                )
                .order_by(Page.page_n)
            ))
            if not pages:
                continue

            referenced: list[str] = []
            for p in pages:
                for tid in (p.table_ids or []):
                    if tid not in referenced:
                        referenced.append(tid)

            tables = []
            if referenced:
                tables = list(db.scalars(
                    select(ExtractedTable).where(ExtractedTable.table_id.in_(referenced))
                ))

            blocks.append(
                f"## Document: {doc.title}  (doc_id={doc.doc_id}, pages {t.start_page}–{t.end_page})"
            )
            for p in pages:
                blocks.append(f"\n### Page {p.page_n}\n{p.prose_text or ''}")
            for tbl in tables:
                blocks.append(f"\n#### Inline table {tbl.table_id} (from page {tbl.source_page})")
                blocks.append(_xlsx_to_markdown(tbl.xlsx_bytes) if tbl.xlsx_bytes else "_(no xlsx data)_")
                inline_table_ids.append(tbl.table_id)

            citations.append(Citation(
                doc_id=doc.doc_id,
                doc_title=doc.title or doc.doc_id,
                pages=f"{t.start_page}-{t.end_page}",
            ))
            blocks.append("\n---\n")

    return "\n".join(blocks), citations, inline_table_ids


def _expand_table_targets_with_pages(routed: RouterResult, window: int = 2) -> None:
    """For each table_target, ensure there's a page_target covering its source page ± window.

    Mutates `routed.page_targets` in place. Prevents the failure mode where the router
    picks a table but no narrative context, leaving the answer agent unable to bridge.
    """
    if not routed.table_targets:
        return
    seen_keys = {(p.doc_id, p.start_page, p.end_page) for p in routed.page_targets}
    extras: list[PageTarget] = []
    with SessionLocal() as db:
        for tt in routed.table_targets:
            tbl = db.get(ExtractedTable, tt.table_id)
            if tbl is None or not tbl.source_page:
                continue
            lo = max(1, tbl.source_page - window)
            hi = tbl.source_page + window
            # Skip if any existing page_target already covers this table's source page
            already_covered = any(
                p.doc_id == tbl.doc_id and p.start_page <= tbl.source_page <= p.end_page
                for p in routed.page_targets
            )
            if already_covered:
                continue
            key = (tbl.doc_id, lo, hi)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            extras.append(PageTarget(
                doc_id=tbl.doc_id,
                start_page=lo,
                end_page=hi,
                reason=f"auto-added narrative context around table {tt.table_id} (page {tbl.source_page})",
            ))
    if extras:
        log.info("[router] auto-added %d page_target(s) around table sources", len(extras))
        for p in extras:
            log.info("    +page_target: %s  p%d-p%d", p.doc_id, p.start_page, p.end_page)
        routed.page_targets.extend(extras)


def _load_table_dataframes(
    table_targets: list[TableTarget],
) -> tuple[dict[str, pd.DataFrame], dict[str, str]]:
    """Load each selected table as a DataFrame for the Excel agent."""
    dfs: dict[str, pd.DataFrame] = {}
    doc_id_by_table: dict[str, str] = {}
    with SessionLocal() as db:
        for t in table_targets:
            tbl = db.get(ExtractedTable, t.table_id)
            if tbl is None or not tbl.xlsx_bytes:
                continue
            try:
                dfs[t.table_id] = _xlsx_to_dataframe(tbl.xlsx_bytes)
                doc_id_by_table[t.table_id] = tbl.doc_id
            except Exception as e:
                log.warning("failed to load table %s as df: %s", t.table_id, e)
    return dfs, doc_id_by_table


# ===========================================================================
# Orchestrator — multi-hop frontier
# ===========================================================================

def _build_router_prompt(
    *,
    question: str,
    original_query: str,
    candidates: list[Doc],
    history: list[HopTrace],
    already_picked_pages: set[str],
    already_picked_tables: set[str],
) -> str:
    """Build a router prompt for one hop. On hop>0 includes history + dedup notice."""
    doc_lines = [
        f"- doc_id={d.doc_id}  title={d.title!r}  pages={d.n_pages}  tables={d.n_tables}\n"
        f"  summary: {(d.doc_summary or '')[:800]}"
        for d in candidates
    ]
    parts = [
        f"User original query: {original_query}",
        f"\nThis hop's investigation question: {question}",
        f"\nCandidate documents ({len(candidates)}):\n" + "\n".join(doc_lines),
    ]
    if history:
        already_pages = "\n".join(f"  - {k}" for k in sorted(already_picked_pages)) or "  (none)"
        already_tables = "\n".join(f"  - {k}" for k in sorted(already_picked_tables)) or "  (none)"
        prior_findings: list[str] = []
        for h in history:
            for f in h.table_findings:
                prior_findings.append(
                    f"  - Table {f.table_id} (doc {f.doc_id}): "
                    f"{f.finding[:200]}{'…' if len(f.finding) > 200 else ''}"
                )
        findings_block = "\n".join(prior_findings) or "  (none)"
        parts.append(
            "\n## Previously explored — DO NOT re-pick these\n"
            f"Pages:\n{already_pages}\n"
            f"Tables:\n{already_tables}\n"
            "\n## Findings from previous hops\n"
            f"{findings_block}\n"
            "\nYour job this hop: address the investigation question above. "
            "Pick DIFFERENT pages/tables that fill the gap. Avoid duplicates."
        )
    else:
        parts.append(
            "\nPick page_targets (and table_targets when tabular reasoning is needed). "
            "Optimise for recall."
        )
    return "\n".join(parts)


def _new_targets(
    routed: RouterResult,
    already_picked_pages: set[str],
    already_picked_tables: set[str],
) -> tuple[list[PageTarget], list[TableTarget]]:
    """Filter router output down to truly new targets (dedup against history)."""
    new_pages: list[PageTarget] = []
    for p in routed.page_targets:
        key = f"{p.doc_id}:{p.start_page}-{p.end_page}"
        if key not in already_picked_pages:
            new_pages.append(p)
            already_picked_pages.add(key)
    new_tables: list[TableTarget] = []
    for t in routed.table_targets:
        if t.table_id not in already_picked_tables:
            new_tables.append(t)
            already_picked_tables.add(t.table_id)
    return new_pages, new_tables


async def _run_excel_on(
    table_targets: list[TableTarget],
    *,
    question: str,
) -> list[TableFinding]:
    """Run the Excel agent on a set of table_targets. Returns the findings."""
    if not table_targets:
        return []
    dfs, doc_by_tbl = _load_table_dataframes(table_targets)
    if not dfs:
        log.warning("excel: no tables loadable as DataFrames")
        return []
    log.info("=" * 60)
    log.info(">>> EXCEL agent  (tables=%d)", len(dfs))
    log.info("=" * 60)
    for tid, df in dfs.items():
        log.info("    loaded df %s  shape=%s  cols=%s",
                 tid, list(df.shape), list(df.columns.astype(str))[:8])
    t = time.monotonic()
    block = "\n".join(
        f"- {tt.table_id}  (doc={tt.doc_id})  reason: {tt.reason}"
        for tt in table_targets if tt.table_id in dfs
    )
    prompt = (
        f"Investigation question: {question}\n\n"
        f"Tables selected by the router:\n{block}\n\n"
        "For each relevant table, call describe_table then run_pandas to extract data. "
        "Return TableFinding entries."
    )
    agent = _build_excel_agent()
    deps = ExcelDeps(dfs=dfs, doc_id_by_table=doc_by_tbl)
    limits = UsageLimits(request_limit=get_settings().agent_excel_request_limit)
    out: ExcelResult = (await agent.run(prompt, deps=deps, usage_limits=limits)).output
    for f in out.findings:
        if not f.doc_id:
            f.doc_id = doc_by_tbl.get(f.table_id, "")
    log.info("<<< excel done in %.1fs — %d finding(s)",
             time.monotonic() - t, len(out.findings))
    for f in out.findings:
        preview = f.finding[:200] + ("…" if len(f.finding) > 200 else "")
        log.info("    finding(%s): %s", f.table_id, preview)
    return out.findings


def _safe_filename(name: str, max_len: int = 80) -> str:
    """Sanitize a filename: lowercase kebab, ASCII only, .md suffix, length-capped."""
    if "/" in name or "\\" in name or ".." in name:
        raise ValueError(f"unsafe filename: {name!r}")
    stem, ext = os.path.splitext(name)
    if ext.lower() != ".md":
        stem = name
    stem = re.sub(r"[^a-zA-Z0-9._-]+", "-", stem.lower()).strip("-_.")
    stem = stem[:max_len] or "answer"
    return f"{stem}.md"


async def _maybe_save_answer(
    *,
    path_to_subdir: str,
    query: str,
    query_id: str,
    answer_result: AnswerResult,
    citations: list[Citation],
    n_hops: int,
    sink: EventSink,
) -> str | None:
    if not answer_result.save_to_file:
        return None

    filename = _safe_filename(answer_result.suggested_filename or f"{query_id}.md")
    out_dir = Path(path_to_subdir/"outputs")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename

    header = (
        f"# {query}\n\n"
        f"_Confidence: {answer_result.confidence} · "
        f"Hops: {n_hops} · Query: `{query_id}`_\n\n"
    )
    body = answer_result.answer.strip() + "\n\n"
    sources = "## Sources\n" + "\n".join(
        f"- {c.doc_title}, pages {c.pages}" for c in citations
    )
    out_path.write_text(header + body + sources, encoding="utf-8")
    await sink.publish_ui(
        "artifact_ready",
        stage="writing_file",
        status="progress",
        message="Saved grounded answer",
        artifacts=[
            file_artifact(
                kind="markdown",
                path=str(out_path),
                filename=out_path.name,
                type="md",
                mime_type="text/markdown",
                bytes=out_path.stat().st_size,
                content=out_path.read_text(encoding="utf-8"),
            )
        ],
    )
    return str(out_path)


async def answer_query(
    workspace_subdir_path: str,
    workspace_id: str,
    user_id: str,
    query: str,
    doc_ids: list[str] | None = None,
    sink: EventSink = EventSink(),
) -> QueryAnswer:
    """End-to-end pipeline: frontier loop of (route → excel → answer)."""
    t0 = time.monotonic()
    s = get_settings()
    stage = "init"

    try:
        # 1. Candidate docs
        with SessionLocal() as db:
            stmt = select(Doc).where(Doc.workspace_id == workspace_id)
            if doc_ids:
                stmt = stmt.where(Doc.doc_id.in_(doc_ids))
            candidates = list(db.scalars(stmt))
        if not candidates:
            raise ValueError(f"No documents found in workspace_id={workspace_id} for the given filter.")
        log.info("answer_query: %d candidate doc(s), query=%r", len(candidates), query)

        query_id = sink.query_id or f"q_{uuid.uuid4().hex[:12]}"
        sink.query_id = query_id
        sink.workspace_id = sink.workspace_id or workspace_id
        sink.agent_type = sink.agent_type or "document_answering"
        await sink.publish_ui(
            "agent_started",
            stage="document_query",
            status="started",
            message="Searching documents",
            data={
                "query": query,
                "user_id": user_id,
                "n_candidate_docs": len(candidates),
            },
        )

        # 2. Frontier state
        pending: deque[str] = deque([query])
        seen_questions: set[str] = {query}
        already_picked_pages: set[str] = set()
        already_picked_tables: set[str] = set()

        acc_page_targets: list[PageTarget] = []
        acc_table_targets: list[TableTarget] = []
        acc_findings: list[TableFinding] = []
        hops: list[HopTrace] = []
        last_ans: AnswerResult | None = None
        last_citations: list[Citation] = []

        router_agent = _build_router()
        answer_agent = _build_answer_agent()
        router_deps = RouterDeps(
            workspace_id=workspace_id,
            user_id=user_id,
            candidate_doc_ids=[d.doc_id for d in candidates],
            already_picked_pages=already_picked_pages,
            already_picked_tables=already_picked_tables,
        )

        # 3. Frontier loop
        while pending and len(hops) < s.agent_max_hops:
            hop_n = len(hops)
            question = pending.popleft()
            log.info("=" * 60)
            log.info(">>> HOP %d  question=%r", hop_n, question)
            log.info("=" * 60)
            await sink.publish_ui(
                "agent_progress",
                stage="document_hop",
                status="progress",
                message="Investigating document question",
                data={"hop": hop_n, "question": question},
            )

            # 3a. Router
            stage = "router"
            log.info(">>> ROUTER  (request_limit=%d)", s.agent_router_request_limit)
            await sink.publish_ui(
                "agent_progress",
                stage="routing",
                status="progress",
                message="Finding relevant document sections",
                data={"hop": hop_n},
            )
            t_r = time.monotonic()
            router_prompt = _build_router_prompt(
                question=question,
                original_query=query,
                candidates=candidates,
                history=hops,
                already_picked_pages=already_picked_pages,
                already_picked_tables=already_picked_tables,
            )
            router_limits = UsageLimits(request_limit=s.agent_router_request_limit)
            routed = (await router_agent.run(
                router_prompt, deps=router_deps, usage_limits=router_limits,
            )).output
            log.info("<<< router done in %.1fs — %d page_target(s), %d table_target(s)",
                     time.monotonic() - t_r, len(routed.page_targets), len(routed.table_targets))
            await sink.publish_ui(
                "agent_progress",
                stage="routing",
                status="progress",
                message="Found relevant document sections",
                data={
                    "hop": hop_n,
                    "page_target_count": len(routed.page_targets),
                    "table_target_count": len(routed.table_targets),
                    "targets": [
                        {
                            "doc_id": p.doc_id,
                            "pages": f"{p.start_page}-{p.end_page}",
                            "reason": p.reason,
                        }
                        for p in routed.page_targets[:5]
                    ],
                    "reasoning": routed.reasoning,
                },
            )
            _expand_table_targets_with_pages(routed)
            new_pages, new_tables = _new_targets(routed, already_picked_pages, already_picked_tables)
            log.info("    new this hop: %d page_target(s), %d table_target(s)",
                     len(new_pages), len(new_tables))
            for pt in new_pages:
                log.info("    +page: %s  p%d-p%d  reason=%r",
                         pt.doc_id, pt.start_page, pt.end_page, pt.reason)
            for tt in new_tables:
                log.info("    +table: %s  (doc=%s)  reason=%r",
                         tt.table_id, tt.doc_id, tt.reason)

            if not new_pages and not new_tables:
                log.info("[hop %d] router emitted no new targets — stopping loop", hop_n)
                break

            acc_page_targets.extend(new_pages)
            acc_table_targets.extend(new_tables)

            # 3b. Excel on NEW tables only
            if new_tables:
                stage = "excel"
                await sink.publish_ui(
                    "agent_progress",
                    stage="excel",
                    status="progress",
                    message="Analyzing extracted tables",
                    data={"n_tables": len(new_tables)},
                )
                new_findings = await _run_excel_on(new_tables, question=question)
                await sink.publish_ui(
                    "artifact_ready",
                    stage="excel",
                    status="progress",
                    message="Table findings are ready",
                    artifacts=[
                        file_artifact(
                            kind="extracted_content",
                            type="json",
                            mime_type="application/json",
                            content=json.dumps([f.model_dump() for f in new_findings], default=str),
                            metadata={"n_tables": len(new_tables)},
                        )
                    ],
                )
            else:
                new_findings = []
            acc_findings.extend(new_findings)

            # 3c. Answer on cumulative context
            context_md, citations, _ = _render_page_context(acc_page_targets)
            findings_block = ""
            if acc_findings:
                findings_block = "\n\n## Tabular analyst findings\n" + "\n".join(
                    f"- Table {f.table_id} (doc {f.doc_id}): {f.finding}"
                    for f in acc_findings
                )
            answer_prompt = (
                f"User query: {query}\n\n"
                f"This hop's investigation question: {question}\n\n"
                f"Provided page content ({len(citations)} section(s)):\n\n{context_md}"
                f"{findings_block}"
            )
            stage = "answer"
            log.info(">>> ANSWER  (context=%d chars, findings=%d)",
                     len(context_md), len(acc_findings))
            await sink.publish_ui(
                "agent_progress",
                stage="answering",
                status="progress",
                message="Drafting grounded answer",
                data={"hop": hop_n},
            )
            t_a = time.monotonic()
            answer_limits = UsageLimits(request_limit=s.agent_answer_request_limit)
            ans = (await answer_agent.run(answer_prompt, usage_limits=answer_limits)).output
            log.info("<<< answer done in %.1fs — confidence=%s  needs_more=%s  follow_ups=%d",
                     time.monotonic() - t_a, ans.confidence, ans.needs_more, len(ans.follow_up_questions))
            await sink.publish_ui(
                "artifact_ready",
                stage="answering",
                status="progress",
                message="Grounded answer draft is ready",
                data={
                    "confidence": ans.confidence,
                    "needs_more": ans.needs_more,
                    "follow_up_questions": ans.follow_up_questions,
                },
                artifacts=[
                    file_artifact(
                        kind="final_answer",
                        type="md",
                        mime_type="text/markdown",
                        content=ans.answer,
                        metadata={"confidence": ans.confidence, "hop": hop_n},
                    )
                ],
            )
            for fq in ans.follow_up_questions:
                log.info("    follow_up: %s", fq)

            last_ans = ans
            last_citations = ans.citations or citations
            hops.append(HopTrace(
                hop=hop_n,
                question=question,
                page_targets=new_pages,
                table_targets=new_tables,
                table_findings=new_findings,
                confidence=ans.confidence,
                needs_more=ans.needs_more,
                follow_up_questions=ans.follow_up_questions,
            ))

            # 3d. Termination checks
            if ans.confidence == "high":
                log.info("[hop %d] confidence=high — stopping loop", hop_n)
                break
            if not ans.needs_more:
                log.info("[hop %d] needs_more=False — stopping loop", hop_n)
                break
            cap = s.agent_max_followups_per_hop
            for fq in ans.follow_up_questions[:cap]:
                if fq not in seen_questions and fq not in pending:
                    seen_questions.add(fq)
                    pending.append(fq)

        # 4. Build the final QueryAnswer
        output_path: str | None = None
        if last_ans is not None:
            output_path = await _maybe_save_answer(
                path_to_subdir=workspace_subdir_path,
                query=query,
                query_id=query_id,
                answer_result=last_ans,
                citations=last_citations,
                n_hops=len(hops),
                sink=sink,
            )

        if last_ans is None:
            result = QueryAnswer(
                query_id=query_id,
                query=query,
                page_targets=[],
                table_targets=[],
                table_findings=[],
                answer="No relevant content was found in the available documents to answer this query.",
                confidence="low",
                citations=[],
                latency_ms=int((time.monotonic() - t0) * 1000),
                n_hops=0,
                hops=[],
                output_path=None,
            )
        else:
            result = QueryAnswer(
                query_id=query_id,
                query=query,
                page_targets=acc_page_targets,
                table_targets=acc_table_targets,
                table_findings=acc_findings,
                answer=last_ans.answer,
                confidence=last_ans.confidence,
                citations=last_citations,
                latency_ms=int((time.monotonic() - t0) * 1000),
                n_hops=len(hops),
                hops=hops,
                output_path=output_path,
            )

        doc_ids_used = sorted(
            {t.doc_id for t in acc_page_targets} | {t.doc_id for t in acc_table_targets}
        )
        table_ids_used = sorted({t.table_id for t in acc_table_targets})
        _save_query(workspace_id, user_id, query, result, doc_ids_used, table_ids_used)
        log.info("answer_query done — hops=%d  confidence=%s  latency=%d ms",
                 result.n_hops, result.confidence, result.latency_ms)
        await sink.publish_ui(
            "agent_ended",
            stage="done",
            status="completed",
            message="Document answer complete",
            data={
                "query_id": result.query_id,
                "confidence": result.confidence,
                "n_hops": result.n_hops,
                "citation_count": len(result.citations),
                "output_path": result.output_path,
            },
        )
        return result

    except Exception as e:
        await sink.publish_ui(
            "agent_ended",
            stage=stage,
            status="failed",
            message="Document answer failed",
            data={"error_class": type(e).__name__, "error": str(e)},
        )
        raise


def _save_query(
    workspace_id: str,
    user_id: str,
    query_text: str,
    result: QueryAnswer,
    doc_ids_used: list[str],
    table_ids_used: list[str],
) -> None:
    with SessionLocal() as db:
        utils.create_query(
            db,
            query_id=result.query_id,
            workspace_id=workspace_id,
            user_id=user_id,
            query_text=query_text,
            doc_ids_used=doc_ids_used,
            table_ids_used=table_ids_used,
            answer=result.answer,
            citations_json=[c.model_dump() for c in result.citations],
            latency_ms=result.latency_ms,
        )
